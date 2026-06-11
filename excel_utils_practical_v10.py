from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from models import Carton, Item
from optimizer_practical_v10 import (
    optimize_one_skid_for_all_items,
    create_failure_visualization_plan,
    print_elapsed,
)
from visualizer_plotly import export_plan_to_plotly_html
from all_skids_visualizer_practical_v10 import export_all_skids_to_plotly_html


DIMENSION_COLS = (1, 2, 3)
GROUP_COL = 4



def safe_filename_part(value: str) -> str:
    """
    Converts a group key like D12:D21 into a safe filename part.
    """

    return (
        str(value)
        .replace(":", "_")
        .replace("/", "_")
        .replace("\\", "_")
        .replace(" ", "_")
    )


@dataclass(frozen=True)
class ExcelGroupInfo:
    """
    Represents one Excel group.

    group_key:
        Unique group identity based on the merged-cell range, not the displayed
        group number. This matters because the same group number can repeat in
        different merged blocks.

    group_value:
        The visible value from column D, usually the top-left value of the
        merged cell.
    """

    group_key: str
    group_value: Any


@dataclass(frozen=True)
class RowItem:
    """
    One carton row from the Excel sheet.
    """

    excel_row_number: int
    original_values: list[Any]
    group_info: ExcelGroupInfo
    item: Item


def parse_positive_float(value: Any, row_number: int, column_number: int) -> float:
    """
    Converts an Excel cell value into a positive float.
    """

    if value is None or str(value).strip() == "":
        raise ValueError(
            f"Row {row_number}, column {column_number}: missing dimension value"
        )

    try:
        number = float(value)
    except ValueError as error:
        raise ValueError(
            f"Row {row_number}, column {column_number}: invalid dimension value {value!r}"
        ) from error

    if number <= 0:
        raise ValueError(
            f"Row {row_number}, column {column_number}: dimension must be greater than 0"
        )

    return number


def is_positive_float(value: Any) -> bool:
    """
    True if value looks like a usable positive number.
    Used for detecting which Excel rows are carton rows.
    """

    try:
        return float(value) > 0
    except (TypeError, ValueError):
        return False


def get_cell_value_with_merged_support(ws, row: int, col: int) -> Any:
    """
    Returns the visible value for a cell.

    For normal cells:
        returns that cell's value.

    For merged cells:
        returns the value from the top-left cell of the merged range.

    This is useful because openpyxl only stores the value on the top-left
    cell of a merged region.
    """

    cell = ws.cell(row=row, column=col)

    if cell.value is not None:
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            return ws.cell(row=min_row, column=min_col).value

    return None


def build_group_lookup(ws) -> dict[int, ExcelGroupInfo]:
    """
    Builds a row -> group mapping for column D.

    Important:
    The key is the merged-cell range, not the number written inside it.

    Example:
        D2:D10 with value 123 becomes group_key="D2:D10"
        D11:D15 with value 123 becomes group_key="D11:D15"

    Even though both values are 123, they are treated as different groups.
    """

    lookup: dict[int, ExcelGroupInfo] = {}

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))

        # Only care about merged ranges that include column D.
        if not (min_col <= GROUP_COL <= max_col):
            continue

        group_value = ws.cell(row=min_row, column=GROUP_COL).value

        # If the merged range starts before column D, the visible value might
        # technically be in the top-left cell of the merged range. Fall back to
        # that if D's top cell is blank.
        if group_value is None:
            group_value = ws.cell(row=min_row, column=min_col).value

        group_key = f"D{min_row}:D{max_row}"

        for row in range(min_row, max_row + 1):
            lookup[row] = ExcelGroupInfo(
                group_key=group_key,
                group_value=group_value,
            )

    return lookup


def get_group_info_for_row(ws, row: int, merged_group_lookup: dict[int, ExcelGroupInfo]) -> ExcelGroupInfo:
    """
    Returns the group info for one row.

    If the row is part of a merged cell in column D, the merged-cell range is
    used as the group identity.

    If it is not merged, that single row becomes its own group. This prevents
    repeated numbers in column D from accidentally being grouped together.
    """

    if row in merged_group_lookup:
        return merged_group_lookup[row]

    group_value = get_cell_value_with_merged_support(ws, row, GROUP_COL)

    return ExcelGroupInfo(
        group_key=f"D{row}:D{row}",
        group_value=group_value,
    )


def detect_data_rows(ws) -> list[int]:
    """
    Detects rows where the first 3 columns are positive numeric dimensions.

    This lets the file work with or without a header row.
    """

    data_rows = []

    for row in range(1, ws.max_row + 1):
        length_value = ws.cell(row=row, column=1).value
        width_value = ws.cell(row=row, column=2).value
        height_value = ws.cell(row=row, column=3).value

        if (
            is_positive_float(length_value)
            and is_positive_float(width_value)
            and is_positive_float(height_value)
        ):
            data_rows.append(row)

    return data_rows


def make_output_headers(ws, first_data_row: int, max_output_col: int) -> list[str]:
    """
    Builds CSV headers.

    If row 1 appears to be a header row, use it.
    Otherwise, create generic headers.
    """

    if first_data_row > 1:
        headers = []

        for col in range(1, max_output_col + 1):
            value = ws.cell(row=1, column=col).value
            headers.append(str(value).strip() if value not in (None, "") else f"column_{col}")

        return headers

    headers = []

    for col in range(1, max_output_col + 1):
        if col == 1:
            headers.append("carton_length")
        elif col == 2:
            headers.append("carton_width")
        elif col == 3:
            headers.append("carton_height")
        elif col == 4:
            headers.append("group_value")
        else:
            headers.append(f"column_{col}")

    return headers


def read_excel_rows(input_path: str, sheet_name: str | None = None) -> tuple[list[RowItem], list[str]]:
    """
    Reads the Excel sheet and returns carton rows with group information.

    The first 3 columns are dimensions.
    The 4th column is the grouped/merged-cell column.
    """

    workbook = load_workbook(input_path, data_only=True)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} was not found. Available sheets: {workbook.sheetnames}"
            )
        ws = workbook[sheet_name]
    else:
        ws = workbook.active

    data_rows = detect_data_rows(ws)

    if not data_rows:
        raise ValueError(
            "No carton rows found. Expected positive numeric dimensions in columns A, B, and C."
        )

    max_output_col = max(ws.max_column, GROUP_COL)
    headers = make_output_headers(ws, first_data_row=min(data_rows), max_output_col=max_output_col)

    merged_group_lookup = build_group_lookup(ws)

    row_items: list[RowItem] = []

    for row in data_rows:
        length = parse_positive_float(ws.cell(row=row, column=1).value, row, 1)
        width = parse_positive_float(ws.cell(row=row, column=2).value, row, 2)
        height = parse_positive_float(ws.cell(row=row, column=3).value, row, 3)

        group_info = get_group_info_for_row(
            ws=ws,
            row=row,
            merged_group_lookup=merged_group_lookup,
        )

        original_values = []

        for col in range(1, max_output_col + 1):
            if col == GROUP_COL:
                # CSV cannot preserve merged cells, so fill the visible group
                # value onto every row in that merged group.
                original_values.append(group_info.group_value)
            else:
                original_values.append(get_cell_value_with_merged_support(ws, row, col))

        item = Item(
            csv_row_number=row,
            copy_number=1,
            carton=Carton(
                length=length,
                width=width,
                height=height,
            ),
        )

        row_items.append(
            RowItem(
                excel_row_number=row,
                original_values=original_values,
                group_info=group_info,
                item=item,
            )
        )

    return row_items, headers


def process_excel_to_csv(
    input_path: str,
    output_path: str,
    sheet_name: str | None = None,
) -> None:
    """
    Main function.

    Reads an Excel sheet with merged groups in column D, optimizes one skid per
    merged group, and writes a CSV with skid dimensions per row.
    """

    print_elapsed(f"Starting Excel processing: {input_path}")

    row_items, headers = read_excel_rows(input_path=input_path, sheet_name=sheet_name)

    groups: dict[str, list[RowItem]] = {}

    for row_item in row_items:
        groups.setdefault(row_item.group_info.group_key, []).append(row_item)

    group_results: dict[str, dict[str, Any]] = {}
    group_plans: dict[str, Any] = {}

    for group_number, (group_key, grouped_rows) in enumerate(groups.items(), start=1):
        print_elapsed(
            f"Starting group {group_number}/{len(groups)} {group_key} "
            f"with {len(grouped_rows)} carton row(s)."
        )

        group_start_time = __import__("time").perf_counter()

        items = [row.item for row in grouped_rows]
        plan = optimize_one_skid_for_all_items(items, debug_label=group_key)

        if plan is None:
            debug_plan = create_failure_visualization_plan(items)

            if debug_plan is not None:
                debug_filename = f"debug_no_valid_skid_{safe_filename_part(group_key)}.html"
                debug_output_path = str(Path(output_path).with_name(debug_filename))
                export_plan_to_plotly_html(debug_plan, debug_output_path)
                group_plans[group_key] = debug_plan
                print(f"Debug visualization for failed group {group_key}: {debug_output_path}")

            group_results[group_key] = {
                "group_range": group_key,
                "group_value": grouped_rows[0].group_info.group_value,
                "cartons_in_group": len(grouped_rows),
                "skid_length": "NO VALID SKID",
                "skid_width": "NO VALID SKID",
                "skid_height": "NO VALID SKID",
            }

            print_elapsed(
                f"Finished group {group_number}/{len(groups)} {group_key}: NO VALID SKID.",
                attempt_start_time=group_start_time,
            )
        else:
            group_results[group_key] = {
                "group_range": group_key,
                "group_value": grouped_rows[0].group_info.group_value,
                "cartons_in_group": len(grouped_rows),
                "skid_length": round(plan.skid_length, 2),
                "skid_width": round(plan.skid_width, 2),
                "skid_height": round(plan.skid_height, 2),
            }

            group_plans[group_key] = plan

            print_elapsed(
                (
                    f"Finished group {group_number}/{len(groups)} {group_key}: "
                    f"VALID skid {round(plan.skid_length, 2)} x "
                    f"{round(plan.skid_width, 2)} x "
                    f"{round(plan.skid_height, 2)}."
                ),
                attempt_start_time=group_start_time,
            )

    output_headers = headers + [
        "merged_group_range",
        "merged_group_value",
        "cartons_in_group",
        "skid_length",
        "skid_width",
        "skid_height",
    ]

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", newline="", encoding="utf-8") as outfile:
        writer = csv.writer(outfile)
        writer.writerow(output_headers)

        for row_item in row_items:
            result = group_results[row_item.group_info.group_key]

            writer.writerow(
                row_item.original_values
                + [
                    result["group_range"],
                    result["group_value"],
                    result["cartons_in_group"],
                    result["skid_length"],
                    result["skid_width"],
                    result["skid_height"],
                ]
            )

    print(f"Processed {len(groups)} group(s).")
    print(f"Processed {len(row_items)} carton row(s).")
    print(f"Wrote output CSV to: {output_path}")
    print_elapsed("Finished writing output CSV.")

    html_start_time = __import__("time").perf_counter()
    all_skids_html_path = str(Path(output_path).with_suffix(".html"))

    export_all_skids_to_plotly_html(
        group_plans=group_plans,
        group_results=group_results,
        output_html_path=all_skids_html_path,
    )

    print_elapsed(
        f"Finished creating all-skids Plotly HTML: {all_skids_html_path}",
        attempt_start_time=html_start_time,
    )
